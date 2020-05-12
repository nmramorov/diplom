from collections import namedtuple
from typing import List
from pprint import pprint

from openpyxl import load_workbook, Workbook


class ExcelDataGetter:
    def __init__(self, filename: str):
        self.excel_file = load_workbook(filename)
        self.filename = filename
        self.required_data = ['index', 'identifier', 'initials', 'team', 'year_of_birth', 'grade', 'first_track',
                              'second_track', 'sum']

    def __repr__(self) -> str:
        return f'Excel file: {self.filename}. Required data is {self.required_data}'

    def get_sheetnames(self) -> List:
        return self.excel_file.sheetnames

    def find_start_values_row(self, sheet: Workbook.active = None) -> int:
        if sheet is None:
            sheet = self.excel_file.active
        for row_index, row_values in enumerate(sheet.iter_rows(min_row=1, max_row=6, values_only=True)):
            if 'Ст.№' in row_values or 'Место' in row_values:
                # Перебор ведется от нулевого ряда, для верного отображения стартового ряда необходимо прибавить 2
                return row_index + 2

    def get_required_data(self, sheet: Workbook.active = None) -> List:
        if sheet is None:
            sheet = self.excel_file.active
        data = []
        start_row = self.find_start_values_row(sheet)
        RequiredDataTemplate = namedtuple('RequiredDataTemplate', self.required_data)
        for row in sheet.iter_rows(min_row=start_row, max_col=9, values_only=True):
            if None in row[1:3]:
                return data
            else:
                data.append(RequiredDataTemplate(*row))

    def get_excel_data(self) -> List:
        return [self.get_required_data(sheet) for sheet in self.excel_file]

    def show_excel_data(self) -> None:
        pprint(self.get_excel_data())
